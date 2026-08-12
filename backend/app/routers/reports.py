import os
import tempfile

from fastapi import APIRouter, Depends
from fastapi.responses import FileResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook

from app.database import get_db
from app.auth import get_current_admin
from app.models import User, Land


router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


def create_excel_response(
    workbook: Workbook,
    filename: str
):
    """
    Save workbook to a temporary file and return it
    as a downloadable Excel response.
    """

    temp_dir = tempfile.gettempdir()
    file_path = os.path.join(temp_dir, filename)

    workbook.save(file_path)

    return FileResponse(
        path=file_path,
        media_type=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        filename=filename,
    )


@router.get("/users")
def export_users(
    db: Session = Depends(get_db),
    admin: int = Depends(get_current_admin)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Users"

    ws.append([
        "ID",
        "Full Name",
        "Email",
        "Mobile",
        "Role"
    ])

    users = (
        db.query(User)
        .order_by(User.id.asc())
        .all()
    )

    for user in users:
        ws.append([
            user.id,
            user.full_name,
            user.email,
            user.mobile,
            user.role
        ])

    return create_excel_response(
        wb,
        "users_report.xlsx"
    )


@router.get("/lands")
def export_lands(
    db: Session = Depends(get_db),
    admin: int = Depends(get_current_admin)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Lands"

    ws.append([
        "ID",
        "Title",
        "Owner ID",
        "Village",
        "Mandal",
        "District",
        "State",
        "Area",
        "Price",
        "Soil Type",
        "Water Source",
        "Crop Type",
        "Status"
    ])

    lands = (
        db.query(Land)
        .order_by(Land.id.asc())
        .all()
    )

    for land in lands:
        ws.append([
            land.id,
            land.title,
            land.owner_id,
            land.village,
            land.mandal,
            land.district,
            land.state,
            land.area,
            land.price,
            land.soil_type,
            land.water_source,
            land.crop_type,
            land.status
        ])

    return create_excel_response(
        wb,
        "lands_report.xlsx"
    )


@router.get("/pending-lands")
def export_pending_lands(
    db: Session = Depends(get_db),
    admin: int = Depends(get_current_admin)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Pending Lands"

    ws.append([
        "ID",
        "Title",
        "Owner ID",
        "Village",
        "Mandal",
        "District",
        "State",
        "Area",
        "Price",
        "Soil Type",
        "Water Source",
        "Crop Type",
        "Status"
    ])

    lands = (
        db.query(Land)
        .filter(Land.status == "pending")
        .order_by(Land.id.asc())
        .all()
    )

    for land in lands:
        ws.append([
            land.id,
            land.title,
            land.owner_id,
            land.village,
            land.mandal,
            land.district,
            land.state,
            land.area,
            land.price,
            land.soil_type,
            land.water_source,
            land.crop_type,
            land.status
        ])

    return create_excel_response(
        wb,
        "pending_lands_report.xlsx"
    )


@router.get("/approved-lands")
def export_approved_lands(
    db: Session = Depends(get_db),
    admin: int = Depends(get_current_admin)
):
    wb = Workbook()
    ws = wb.active
    ws.title = "Approved Lands"

    ws.append([
        "ID",
        "Title",
        "Owner ID",
        "Village",
        "Mandal",
        "District",
        "State",
        "Area",
        "Price",
        "Soil Type",
        "Water Source",
        "Crop Type",
        "Status"
    ])

    lands = (
        db.query(Land)
        .filter(Land.status == "approved")
        .order_by(Land.id.asc())
        .all()
    )

    for land in lands:
        ws.append([
            land.id,
            land.title,
            land.owner_id,
            land.village,
            land.mandal,
            land.district,
            land.state,
            land.area,
            land.price,
            land.soil_type,
            land.water_source,
            land.crop_type,
            land.status
        ])

    return create_excel_response(
        wb,
        "approved_lands_report.xlsx"
    )